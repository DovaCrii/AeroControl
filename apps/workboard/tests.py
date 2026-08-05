from datetime import timedelta
import pytest
from django.contrib.auth.models import Permission, User
from rest_framework.authtoken.models import Token
from django.test import Client
from django.urls import reverse
from django.utils import timezone

from apps.registry.models import CostCenter, Operator
from apps.core.models import OperationalTenant, TenantMembership
from .models import (
    KanbanBoard,
    KanbanBoardAccess,
    KanbanChecklistItem,
    KanbanLabel,
    KanbanStage,
    KanbanTask,
)


@pytest.fixture
def user(db):
    user = User.objects.create_user("operator", password="password")
    user.user_permissions.add(
        *Permission.objects.filter(
            content_type__app_label="workboard",
            content_type__model__in=[
                "kanbantask",
                "kanbanstage",
                "kanbanboard",
                "kanbanchecklistitem",
                "kanbanlabel",
            ],
            codename__in=[
                "add_kanbantask",
                "change_kanbantask",
                "view_kanbantask",
                "add_kanbanstage",
                "add_kanbanboard",
                "change_kanbanboard",
                "add_kanbanchecklistitem",
                "change_kanbanchecklistitem",
                "add_kanbanlabel",
                "view_kanbanlabel",
            ],
        )
    )
    return user


@pytest.fixture
def auth_client(user):
    client = Client()
    assert client.login(username="operator", password="password")
    return client


@pytest.fixture
def board(db):
    board = KanbanBoard.objects.create(name="Operations")
    todo = KanbanStage.objects.create(board=board, name="Todo", order=0)
    done = KanbanStage.objects.create(board=board, name="Done", order=1)
    return board, todo, done


@pytest.fixture
def operator(db):
    cost_center = CostCenter.objects.create(code="OPS", name="Operations")
    return Operator.objects.create(
        employee_id="EMP-001",
        full_name="Test Operator",
        cost_center=cost_center,
    )


@pytest.mark.django_db
def test_kanban_and_mutating_endpoints_require_auth(board):
    _, todo, _ = board
    task = KanbanTask.objects.create(board=todo.board, stage=todo, title="Task")
    client = Client()

    assert client.get(reverse("kanban")).status_code == 302
    assert (
        client.post(
            reverse("task-move", args=[task.pk]), {"stage_id": todo.pk}
        ).status_code
        == 302
    )
    assert client.get(reverse("task-quick"), {"stage_id": todo.pk}).status_code == 302
    assert (
        client.post(
            reverse("task-quick"), {"stage_id": todo.pk, "title": "Task"}
        ).status_code
        == 302
    )


@pytest.mark.django_db
def test_authenticated_kanban_renders_and_empty_board_state(auth_client, board):
    board_obj, _, _ = board
    response = auth_client.get(reverse("kanban"))
    assert response.status_code == 200
    assert "Operations" in response.content.decode()

    board_obj.stages.update(is_active=False)
    response = auth_client.get(reverse("kanban"))
    assert response.status_code == 200
    assert "Este tablero no tiene etapas activas." in response.content.decode()


@pytest.mark.django_db
def test_kanban_without_boards_shows_empty_state(auth_client):
    response = auth_client.get(reverse("kanban"))
    assert response.status_code == 200
    assert "No hay ningún tablero configurado." in response.content.decode()


@pytest.mark.django_db
def test_kanban_filters_by_operator_and_priority(auth_client, board, operator):
    board_obj, todo, _ = board
    matching = KanbanTask.objects.create(
        board=board_obj,
        stage=todo,
        title="Matching",
        assigned_to=operator,
        priority="high",
    )
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="Other", priority="low"
    )

    response = auth_client.get(
        reverse("kanban"), {"operator": operator.pk, "priority": "high"}
    )
    content = response.content.decode()
    assert response.status_code == 200
    assert "Matching" in content
    assert "Other" not in content
    assert 'data-drag-disabled="true"' in content
    assert "El orden por arrastre está desactivado" in content
    assert str(matching.pk) in content


@pytest.mark.django_db
def test_malformed_board_and_operator_filters_are_ignored(auth_client, board):
    response = auth_client.get(
        reverse("kanban"), {"board": "not-a-uuid", "operator": "also-not-a-uuid"}
    )
    assert response.status_code == 200
    assert "Operations" in response.content.decode()

    response = auth_client.get(
        reverse("kanban-board-partial"),
        {"board": "not-a-uuid", "operator": "also-not-a-uuid"},
    )
    assert response.status_code == 302
    assert response.url.startswith(reverse("kanban"))

    response = auth_client.get(
        reverse("kanban-board-partial"),
        {"board": board[0].pk, "operator": "also-not-a-uuid"},
        HTTP_HX_REQUEST="true",
    )
    assert response.status_code == 200
    assert response.headers["HX-Push-Url"].startswith(reverse("kanban"))


@pytest.mark.django_db
def test_stage_create_is_available_from_empty_board(auth_client, board):
    board_obj, _, _ = board
    response = auth_client.post(
        reverse("stage-create") + f"?board={board_obj.pk}",
        {"board": board_obj.pk, "name": "Review", "order": 2, "color": "#2EC4B6"},
    )

    assert response.status_code == 302
    assert response.url == f"{reverse('kanban')}?board={board_obj.pk}"
    assert board_obj.stages.filter(name="Review").exists()


def test_order_is_not_a_kanban_form_field():
    """`order` is a technical position managed by drag-and-drop, not typed."""
    from apps.workboard.forms import (
        KanbanChecklistItemForm,
        KanbanLabelForm,
        KanbanStageForm,
        KanbanTaskForm,
    )

    for form_cls in (
        KanbanStageForm,
        KanbanTaskForm,
        KanbanLabelForm,
        KanbanChecklistItemForm,
    ):
        assert "order" not in form_cls().fields


@pytest.mark.django_db
def test_new_stage_label_and_task_are_appended_at_the_end(auth_client, board):
    """`order` left the forms; the create views assign it server-side so a new
    column/label/card lands at the end instead of jumping to the front."""
    board_obj, todo, _ = board  # the fixture seeds two stages (order 0 and 1)

    stage_resp = auth_client.post(
        reverse("stage-create") + f"?board={board_obj.pk}",
        {"board": board_obj.pk, "name": "Review", "color": "#2EC4B6"},
    )
    assert stage_resp.status_code == 302
    assert KanbanStage.objects.get(board=board_obj, name="Review").order == 2

    KanbanLabel.objects.create(board=board_obj, name="Urgent", color="#ff0000", order=0)
    label_resp = auth_client.post(
        reverse("label-create"),
        {"board": board_obj.pk, "name": "Later", "color": "#0000ff"},
    )
    assert label_resp.status_code == 302
    assert KanbanLabel.objects.get(board=board_obj, name="Later").order == 1

    KanbanTask.objects.create(board=board_obj, stage=todo, title="First", order=0)
    task_resp = auth_client.post(
        reverse("task-create"),
        {"board": board_obj.pk, "stage": todo.pk, "title": "Second", "priority": "low"},
    )
    assert task_resp.status_code == 302
    assert KanbanTask.objects.get(board=board_obj, title="Second").order == 1


@pytest.mark.django_db
def test_board_and_task_archives_are_reversible(auth_client, board):
    board_obj, todo, _ = board
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Archive me")

    task_response = auth_client.post(reverse("task-archive", args=[task.pk]))
    assert task_response.status_code == 302
    task.refresh_from_db()
    assert task.is_active is False

    board_response = auth_client.post(reverse("board-archive", args=[board_obj.pk]))
    assert board_response.status_code == 302
    board_obj.refresh_from_db()
    assert board_obj.is_active is False


@pytest.mark.django_db
def test_quick_add_preserves_filters_and_refreshes_column(auth_client, board, operator):
    board_obj, todo, _ = board
    KanbanTask.objects.create(
        board=board_obj,
        stage=todo,
        title="Visible",
        assigned_to=operator,
        priority="high",
    )
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="Hidden", priority="low"
    )

    form = auth_client.get(
        reverse("task-quick"),
        {
            "stage_id": todo.pk,
            "board": board_obj.pk,
            "operator": operator.pk,
            "priority": "high",
        },
    )
    assert form.status_code == 200
    form_content = form.content.decode()
    assert f'name="board" value="{board_obj.pk}"' in form_content
    assert f'name="operator" value="{operator.pk}"' in form_content
    assert 'name="filter_priority" value="high"' in form_content
    assert '<option value="high" selected>' not in form_content
    assert '<option value="medium" selected>Medium</option>' in form_content

    response = auth_client.post(
        reverse("task-quick"),
        {
            "stage_id": todo.pk,
            "title": "New visible",
            "assigned_to": operator.pk,
            "priority": "high",
            "filter_priority": "high",
            "board": board_obj.pk,
            "operator": operator.pk,
        },
    )
    assert response.status_code == 200
    content = response.content.decode()
    assert "Visible" in content and "New visible" in content
    assert "Hidden" not in content
    assert "&amp;board=" in content
    assert "&amp;operator=" in content
    assert "&amp;priority=high" in content
    assert "kanban-column" in content
    assert "HX-Retarget" not in response
    assert "HX-Reswap" not in response


@pytest.mark.django_db
def test_quick_add_rejects_stage_on_inactive_board(auth_client, board):
    _, todo, _ = board
    todo.board.is_active = False
    todo.board.save(update_fields=["is_active"])
    response = auth_client.post(
        reverse("task-quick"), {"stage_id": todo.pk, "title": "Invalid"}
    )
    assert response.status_code == 400
    assert not KanbanTask.objects.filter(title="Invalid").exists()


@pytest.mark.django_db
def test_move_updates_stage_and_order_and_rejects_invalid_stage(auth_client, board):
    board_obj, todo, done = board
    first = KanbanTask.objects.create(
        board=board_obj, stage=todo, title="First", order=0
    )
    second = KanbanTask.objects.create(
        board=board_obj, stage=todo, title="Second", order=1
    )

    response = auth_client.post(
        reverse("task-move", args=[second.pk]), {"stage_id": done.pk, "new_order": 0}
    )
    assert response.status_code == 204
    second.refresh_from_db()
    first.refresh_from_db()
    assert second.stage_id == done.pk
    assert second.order == 0
    assert first.order == 0

    other_board = KanbanBoard.objects.create(name="Other")
    other_stage = KanbanStage.objects.create(board=other_board, name="Other stage")
    response = auth_client.post(
        reverse("task-move", args=[first.pk]),
        {"stage_id": other_stage.pk, "new_order": 0},
    )
    assert response.status_code == 400

    done.is_active = False
    done.save(update_fields=["is_active"])
    response = auth_client.post(
        reverse("task-move", args=[first.pk]), {"stage_id": done.pk, "new_order": 0}
    )
    assert response.status_code == 400
    assert first.stage_id == todo.pk

    response = auth_client.post(
        reverse("task-move", args=[first.pk]),
        {"stage_id": "not-a-uuid", "new_order": 0},
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_move_rejects_task_from_inactive_board(auth_client, board):
    board_obj, todo, done = board
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Task")
    board_obj.is_active = False
    board_obj.save(update_fields=["is_active"])
    response = auth_client.post(
        reverse("task-move", args=[task.pk]), {"stage_id": done.pk, "new_order": 0}
    )
    assert response.status_code == 404


@pytest.mark.django_db
def test_move_rejects_task_with_mismatched_board_and_stage(auth_client, board):
    board_obj, todo, done = board
    other_board = KanbanBoard.objects.create(name="Other")
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Task")
    task.board = other_board
    task.save(update_fields=["board"])

    response = auth_client.post(
        reverse("task-move", args=[task.pk]), {"stage_id": done.pk, "new_order": 0}
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_move_requires_csrf(auth_client, board):
    _, todo, done = board
    task = KanbanTask.objects.create(board=todo.board, stage=todo, title="Task")
    client = Client(enforce_csrf_checks=True)
    assert client.login(username="operator", password="password")

    response = client.post(
        reverse("task-move", args=[task.pk]), {"stage_id": done.pk, "new_order": 0}
    )
    assert response.status_code == 403


@pytest.mark.django_db
def test_quick_add_validates_active_assignee(auth_client, board, operator):
    _, todo, _ = board
    inactive = Operator.objects.create(
        employee_id="EMP-002",
        full_name="Inactive Operator",
        cost_center=operator.cost_center,
        is_active=False,
    )
    url = reverse("task-quick")

    for assigned_to in ("not-a-uuid", inactive.pk):
        response = auth_client.post(
            url, {"stage_id": todo.pk, "title": "Invalid", "assigned_to": assigned_to}
        )
        assert response.status_code == 400
    assert not KanbanTask.objects.filter(title="Invalid").exists()


@pytest.mark.django_db
def test_quick_add_creates_task_and_renders_assignee(auth_client, board, operator):
    board_obj, todo, _ = board
    response = auth_client.post(
        reverse("task-quick"),
        {
            "stage_id": todo.pk,
            "title": "Inspect aircraft",
            "priority": "high",
            "assigned_to": operator.pk,
        },
    )
    assert response.status_code == 200
    task = KanbanTask.objects.get(title="Inspect aircraft")
    assert task.board_id == board_obj.pk
    assert task.assigned_to_id == operator.pk
    assert task.priority == "high"

    form = auth_client.get(reverse("task-quick"), {"stage_id": todo.pk})
    assert "Test Operator" in form.content.decode()


@pytest.mark.django_db
def test_workboard_urls_keep_kanban_and_task_list_distinct(auth_client):
    assert reverse("kanban") == "/workboard/"
    assert reverse("task-list") == "/workboard/tasks/"
    assert auth_client.get("/workboard/").status_code == 200
    assert auth_client.get("/workboard/tasks/").status_code == 200


@pytest.mark.django_db
def test_task_detail_checklist_progress_and_list_filters(auth_client, board):
    board_obj, todo, _ = board
    label = KanbanLabel.objects.create(board=board_obj, name="Safety", color="#EF4444")
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Inspect")
    task.labels.add(label)
    first = KanbanChecklistItem.objects.create(task=task, title="Review log", order=0)
    KanbanChecklistItem.objects.create(
        task=task, title="Sign off", order=1, is_completed=True
    )

    detail = auth_client.get(reverse("task-detail", args=[task.pk]))
    assert detail.status_code == 200
    assert "50%" in detail.content.decode()

    response = auth_client.post(reverse("checklist-toggle", args=[first.pk]))
    assert response.status_code == 200
    task.refresh_from_db()
    assert task.checklist_progress == 100

    listing = auth_client.get(
        reverse("workboard-list"), {"q": "Inspect", "label": label.pk}
    )
    assert listing.status_code == 200
    assert "Inspect" in listing.content.decode()


@pytest.mark.django_db
def test_task_report_exports_filtered_csv(auth_client, board):
    board_obj, todo, _ = board
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="Export me", priority="high"
    )
    response = auth_client.get(reverse("task-report-csv"), {"priority": "high"})
    assert response.status_code == 200
    assert "Export me" in response.content.decode()
    assert response["Content-Disposition"].endswith('aerocontrol-tasks.csv"')


@pytest.mark.django_db
def test_task_report_exports_xlsx(auth_client, board):
    board_obj, todo, _ = board
    KanbanTask.objects.create(board=board_obj, stage=todo, title="XLSX task")
    response = auth_client.get(reverse("task-report-xlsx"))
    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/vnd.openxmlformats")
    assert response.content[:2] == b"PK"


@pytest.mark.django_db
def test_task_report_exports_docx(auth_client, board):
    board_obj, todo, _ = board
    KanbanTask.objects.create(board=board_obj, stage=todo, title="Word task")
    response = auth_client.get(reverse("task-report-docx"))
    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/vnd.openxmlformats")
    assert response.content[:2] == b"PK"


@pytest.mark.django_db
def test_csv_report_neutralizes_formula_titles(auth_client, board):
    """T4.4: a task title starting with a formula character must be inert in
    the export (Excel/LibreOffice execute a cell that starts with =/+/-/@)."""
    board_obj, todo, _ = board
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title='=HYPERLINK("http://evil","x")'
    )
    response = auth_client.get(reverse("task-report-csv"))
    body = response.content.decode()
    assert "'=HYPERLINK" in body  # apostrophe-prefixed, inert


@pytest.mark.django_db
def test_xlsx_report_neutralizes_formula_titles(auth_client, board):
    from io import BytesIO

    from openpyxl import load_workbook

    board_obj, todo, _ = board
    KanbanTask.objects.create(board=board_obj, stage=todo, title="=SUM(A1:A2)")
    response = auth_client.get(reverse("task-report-xlsx"))
    sheet = load_workbook(BytesIO(response.content)).active
    # Row 2 (after the header), first column (Task).
    assert sheet.cell(row=2, column=1).value == "'=SUM(A1:A2)"


@pytest.mark.django_db
def test_api_v1_tasks_is_permissioned_and_paginated(auth_client, board):
    board_obj, todo, _ = board
    KanbanTask.objects.create(board=board_obj, stage=todo, title="API task")
    response = auth_client.get("/api/v1/workboard/tasks/", {"page_size": 1})
    assert response.status_code == 200
    payload = response.json()
    assert payload["version"] == "v1"
    assert payload["page_size"] == 1
    assert payload["results"][0]["title"] == "API task"


@pytest.mark.django_db
def test_api_v1_tasks_requires_auth(board):
    response = Client().get("/api/v1/workboard/tasks/")
    assert response.status_code == 401
    assert response.json()["detail"] == "Authentication required."


@pytest.mark.django_db
def test_api_v1_index_describes_contract(auth_client):
    response = auth_client.get("/api/v1/")
    assert response.status_code == 200
    payload = response.json()
    assert payload["version"] == "v1"
    assert payload["endpoints"]["task_update"]["method"] == "PATCH"


@pytest.mark.django_db
def test_api_v1_task_patch_validates_and_updates(auth_client, board):
    board_obj, todo, done = board
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Before")
    response = auth_client.patch(
        f"/api/v1/workboard/tasks/{task.pk}/",
        data='{"title":"After","stage_id":"%s","priority":"high"}' % done.pk,
        content_type="application/json",
    )
    assert response.status_code == 200
    task.refresh_from_db()
    assert task.title == "After"
    assert task.stage_id == done.pk

    invalid = auth_client.patch(
        f"/api/v1/workboard/tasks/{task.pk}/",
        data='{"priority":"invalid"}',
        content_type="application/json",
    )
    assert invalid.status_code == 400

    task.refresh_from_db()
    stale = (task.updated_at - timedelta(microseconds=1)).isoformat()
    task.title = "Changed elsewhere"
    task.save()
    conflict = auth_client.patch(
        f"/api/v1/workboard/tasks/{task.pk}/",
        data='{"title":"Stale update"}',
        content_type="application/json",
        HTTP_IF_UNMODIFIED_SINCE=stale,
    )
    assert conflict.status_code == 409


@pytest.mark.django_db
def test_api_v1_task_patch_records_task_audit_context(auth_client, board):
    from apps.core.models import AuditEvent

    board_obj, todo, done = board
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Before")
    response = auth_client.patch(
        f"/api/v1/workboard/tasks/{task.pk}/",
        data='{"title":"After"}',
        content_type="application/json",
    )
    assert response.status_code == 200
    # set_audit_context runs on the DRF request; the middleware must still see
    # the task context on the underlying HttpRequest, not a generic patch_success.
    event = AuditEvent.objects.get(path=f"/api/v1/workboard/tasks/{task.pk}/")
    assert event.model_label == KanbanTask._meta.label
    assert event.object_id == str(task.pk)


@pytest.mark.django_db
def test_board_object_access_scopes_api(user, auth_client, board):
    board_obj, todo, _ = board
    other = KanbanBoard.objects.create(name="Restricted")
    other_stage = KanbanStage.objects.create(board=other, name="Todo")
    KanbanTask.objects.create(board=board_obj, stage=todo, title="Allowed")
    KanbanTask.objects.create(board=other, stage=other_stage, title="Hidden")
    KanbanBoardAccess.objects.create(board=board_obj, user=user, role="viewer")
    response = auth_client.get("/api/v1/workboard/tasks/")
    assert response.status_code == 200
    titles = [item["title"] for item in response.json()["results"]]
    assert titles == ["Allowed"]


@pytest.mark.django_db
def test_tenant_membership_scopes_api(user, auth_client, board):
    board_obj, todo, _ = board
    tenant = OperationalTenant.objects.create(name="Tenant A", slug="tenant-a")
    TenantMembership.objects.create(tenant=tenant, user=user, role="member")
    board_obj.tenant = tenant
    board_obj.save(update_fields=["tenant", "updated_at"])
    KanbanTask.objects.create(board=board_obj, stage=todo, title="Tenant task")
    other_board = KanbanBoard.objects.create(name="Tenant B")
    other_stage = KanbanStage.objects.create(board=other_board, name="Todo")
    KanbanTask.objects.create(board=other_board, stage=other_stage, title="Shared task")
    response = auth_client.get("/api/v1/workboard/tasks/")
    assert response.status_code == 200
    titles = [item["title"] for item in response.json()["results"]]
    assert "Tenant task" in titles
    assert "Shared task" in titles


@pytest.mark.django_db
def test_drf_api_tasks_is_permissioned(auth_client, board):
    board_obj, todo, _ = board
    KanbanTask.objects.create(board=board_obj, stage=todo, title="DRF task")
    response = auth_client.get("/api/drf/v1/workboard/tasks/")
    assert response.status_code == 200
    assert response.json()[0]["title"] == "DRF task"


@pytest.mark.django_db
def test_drf_api_tasks_requires_auth(board):
    response = Client().get("/api/drf/v1/workboard/tasks/")
    assert response.status_code == 401
    assert response.json()["detail"] == "Authentication required."


@pytest.mark.django_db
def test_drf_api_tasks_requires_view_permission(board):
    client = Client()
    User.objects.create_user("no-task-view", password="password")
    assert client.login(username="no-task-view", password="password")

    response = client.get("/api/drf/v1/workboard/tasks/")

    assert response.status_code == 403


@pytest.mark.django_db
def test_drf_api_tasks_scopes_tenant_and_board_access(user, auth_client, board):
    board_obj, todo, _ = board
    tenant = OperationalTenant.objects.create(name="Tenant A", slug="tenant-a")
    TenantMembership.objects.create(tenant=tenant, user=user, role="member")
    board_obj.tenant = tenant
    board_obj.save(update_fields=["tenant", "updated_at"])
    KanbanTask.objects.create(board=board_obj, stage=todo, title="Tenant task")
    other_tenant = OperationalTenant.objects.create(name="Tenant B", slug="tenant-b")
    other = KanbanBoard.objects.create(name="Tenant B", tenant=other_tenant)
    other_stage = KanbanStage.objects.create(board=other, name="Todo")
    KanbanTask.objects.create(board=other, stage=other_stage, title="Hidden task")

    response = auth_client.get("/api/drf/v1/workboard/tasks/")

    assert response.status_code == 200
    titles = [item["title"] for item in response.json()]
    assert "Tenant task" in titles
    assert "Hidden task" not in titles


@pytest.mark.django_db
def test_api_token_and_openapi_contract(user, auth_client):
    token_response = Client().post(
        "/api-token/", {"username": "operator", "password": "password"}
    )
    assert token_response.status_code == 200
    token = token_response.json()["token"]
    assert token == Token.objects.get(user=user).key

    schema_response = Client().get("/api/v1/openapi.json")
    assert schema_response.status_code == 200
    schema = schema_response.json()
    assert schema["openapi"] == "3.0.3"
    token_auth = schema["components"]["securitySchemes"]["tokenAuth"]
    assert token_auth["type"] == "apiKey"
    assert token_auth["in"] == "header"
    assert token_auth["name"] == "Authorization"
    assert "/api/drf/v1/workboard/tasks/" in schema["paths"]


class TestBoardScopeEnforcement:
    """V.1/V.2/V.5: read and write paths that skipped board scoping.

    The scenario everywhere: the user holds the standard model permissions
    (Operations role) but the board declares access rules where they are only
    a viewer -- or no rule at all, which excludes them once rules exist.
    """

    @staticmethod
    def _restricted_world(user):
        mine = KanbanBoard.objects.create(name="Mine")
        mine_stage = KanbanStage.objects.create(board=mine, name="Todo", order=0)
        other = KanbanBoard.objects.create(name="Theirs")
        other_stage = KanbanStage.objects.create(board=other, name="Todo", order=0)
        KanbanBoardAccess.objects.create(board=mine, user=user, role="viewer")
        mine_task = KanbanTask.objects.create(
            board=mine, stage=mine_stage, title="Visible task"
        )
        other_task = KanbanTask.objects.create(
            board=other, stage=other_stage, title="Hidden task"
        )
        return mine, mine_task, other, other_stage, other_task

    @pytest.mark.django_db
    def test_csv_export_only_ships_visible_tasks(self, auth_client, user):
        self._restricted_world(user)

        response = auth_client.get(reverse("task-list"), {"export": "csv"})
        # The export streams now, so the body is assembled from chunks.
        content = b"".join(
            chunk.encode() if isinstance(chunk, str) else chunk
            for chunk in response.streaming_content
        ).decode("utf-8-sig")

        assert response.status_code == 200
        assert "Visible task" in content
        # Before WList.get_queryset existed, this exported every tenant's rows.
        assert "Hidden task" not in content

    @pytest.mark.django_db
    def test_html_list_is_scoped_too(self, auth_client, user):
        self._restricted_world(user)

        content = auth_client.get(reverse("task-list")).content.decode()

        assert "Visible task" in content
        assert "Hidden task" not in content

    @pytest.mark.django_db
    def test_board_viewer_cannot_edit_a_task(self, auth_client, user):
        _, task, _, _, _ = self._restricted_world(user)

        response = auth_client.post(
            reverse("task-edit", args=[task.pk]),
            {
                "board": task.board_id,
                "stage": task.stage_id,
                "title": "Renamed",
                "priority": "low",
                "order": 0,
            },
        )

        task.refresh_from_db()
        assert response.status_code == 403
        assert task.title == "Visible task"

    @pytest.mark.django_db
    def test_task_cannot_be_moved_to_an_inaccessible_board(self, auth_client, user):
        mine, task, other, other_stage, _ = self._restricted_world(user)
        # Editor on their own board, so the edit itself is allowed.
        access = KanbanBoardAccess.objects.get(board=mine, user=user)
        access.role = "editor"
        access.save(update_fields=["role"])

        response = auth_client.post(
            reverse("task-edit", args=[task.pk]),
            {
                "board": other.pk,
                "stage": other_stage.pk,
                "title": "Kidnapped",
                "priority": "low",
                "order": 0,
            },
        )

        task.refresh_from_db()
        assert task.board_id == mine.pk, response.status_code

    @pytest.mark.django_db
    def test_board_viewer_cannot_create_a_stage(self, auth_client, user):
        mine, _, _, _, _ = self._restricted_world(user)

        response = auth_client.post(
            reverse("stage-create"),
            {
                "board": mine.pk,
                "name": "Sneaky",
                "order": 9,
                "color": "#123456",
                "status_type": "pending",
            },
        )

        assert response.status_code == 403
        assert not KanbanStage.objects.filter(name="Sneaky").exists()

    @pytest.mark.django_db
    def test_board_viewer_cannot_touch_checklists(self, auth_client, user):
        _, task, _, _, _ = self._restricted_world(user)
        item = KanbanChecklistItem.objects.create(task=task, title="Step", order=0)

        created = auth_client.post(
            reverse("checklist-create", args=[task.pk]), {"title": "New", "order": 1}
        )
        toggled = auth_client.post(reverse("checklist-toggle", args=[item.pk]))

        item.refresh_from_db()
        assert created.status_code == 403
        assert toggled.status_code == 403
        assert item.is_completed is False


@pytest.mark.django_db
def test_api_token_endpoint_throttles_credential_guessing(db):
    from django.core.cache import cache

    cache.clear()  # the throttle history lives in the cache
    client = Client()

    statuses = [
        client.post(
            reverse("api-token"), {"username": "ghost", "password": f"try-{i}"}
        ).status_code
        for i in range(12)
    ]

    # The first attempts fail with 400 (bad credentials); the throttle must
    # cut in before the 12th. Without it this endpoint was an unlimited
    # password oracle.
    assert 429 in statuses
    cache.clear()


class TestBoardRenderQueryBudget:
    """V.14/V.15: the board partial re-renders on every drag and filter, so its
    query count must stay flat as tasks and checklists grow."""

    @pytest.mark.django_db
    def test_stage_data_query_count_is_independent_of_cards(
        self, django_assert_max_num_queries
    ):
        from apps.workboard.selectors import build_stage_data

        board = KanbanBoard.objects.create(name="Busy")
        stages = [
            KanbanStage.objects.create(board=board, name=f"S{i}", order=i)
            for i in range(6)
        ]
        for i in range(30):
            task = KanbanTask.objects.create(
                board=board, stage=stages[i % 6], title=f"T{i}", order=i
            )
            for j in range(3):
                KanbanChecklistItem.objects.create(
                    task=task, title=f"step {j}", order=j, is_completed=j == 0
                )

        # 1 stages + 1 tasks + 2 prefetches; headroom for session noise. The
        # old shape was ~3 per stage plus one COUNT per card (~48 here).
        with django_assert_max_num_queries(6):
            data = build_stage_data(board, {})
            for column in data:
                for task in column["tasks"]:
                    assert task.checklist_progress >= 0

        assert sum(len(column["tasks"]) for column in data) == 30


@pytest.mark.django_db
def test_api_patch_validates_values_before_saving(board):
    """V.20: setattr+save without full_clean turned a malformed date into an
    unhandled 500 and persisted oversized titles that PostgreSQL would reject."""
    board_obj, todo, _ = board
    task = KanbanTask.objects.create(board=board_obj, stage=todo, title="Valid")
    user = User.objects.create_user("api", password="password")
    user.user_permissions.add(
        *Permission.objects.filter(
            codename__in=["view_kanbantask", "change_kanbantask"]
        )
    )
    token = Token.objects.create(user=user)
    client = Client(HTTP_AUTHORIZATION=f"Token {token.key}")
    url = reverse("api-v1-workboard-task-update", args=[task.pk])

    bad_date = client.patch(
        url, '{"due_date": "manana"}', content_type="application/json"
    )
    long_title = client.patch(
        url, '{"title": "' + "x" * 300 + '"}', content_type="application/json"
    )

    task.refresh_from_db()
    assert bad_date.status_code == 400
    assert long_title.status_code == 400
    assert task.title == "Valid"


@pytest.mark.parametrize(
    ("offset", "expected"),
    [
        (-1, "overdue"),
        (0, "due_7"),
        (7, "due_7"),
        (8, "due_15"),
        (15, "due_15"),
        (16, "due_30"),
        (30, "due_30"),
        (31, ""),
        (None, ""),
    ],
)
def test_urgency_bucket_matches_the_compliance_report_boundaries(offset, expected):
    """B3.3: same expired/due_7/due_15/due_30 boundaries as
    apps.compliance.reports._vigencia_bucket_counts, so "urgent" means the
    same thing everywhere in the app."""
    today = timezone.localdate()
    task = KanbanTask(
        due_date=None if offset is None else today + timedelta(days=offset)
    )

    assert task.urgency_bucket(today) == expected


@pytest.mark.django_db
def test_kanban_card_shows_graduated_urgency_not_colour_alone(auth_client, board):
    """B3.3: each bucket appends its own translated label so the signal
    survives without colour (screen reader, colour-blind, printed page)."""
    board_obj, todo, _ = board
    today = timezone.localdate()
    KanbanTask.objects.create(
        board=board_obj,
        stage=todo,
        title="Overdue task",
        due_date=today - timedelta(days=1),
    )
    KanbanTask.objects.create(
        board=board_obj,
        stage=todo,
        title="Due soon task",
        due_date=today + timedelta(days=5),
    )
    KanbanTask.objects.create(
        board=board_obj,
        stage=todo,
        title="Far out task",
        due_date=today + timedelta(days=60),
    )

    content = auth_client.get(reverse("kanban")).content.decode()

    assert "is-overdue" in content
    assert "text-warning-emphasis" in content
    assert "Due within 7 days" in content or "Vence en 7 días" in content


@pytest.mark.django_db
def test_kanban_column_header_shows_overdue_count_alongside_total(auth_client, board):
    """B3.4: a loaded column ("47 tasks") also surfaces how many are actually
    late, without opening every card."""
    board_obj, todo, done = board
    today = timezone.localdate()
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="Late", due_date=today - timedelta(days=2)
    )
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="On time", due_date=today + timedelta(days=2)
    )
    KanbanTask.objects.create(board=board_obj, stage=done, title="Done, no due date")

    content = auth_client.get(reverse("kanban")).content.decode()

    assert "kanban-count-overdue" in content
    assert "Overdue in this stage: 1" in content or "esta etapa: 1" in content


@pytest.mark.django_db
def test_kanban_column_header_hides_overdue_badge_when_nothing_is_late(
    auth_client, board
):
    board_obj, todo, _ = board
    today = timezone.localdate()
    KanbanTask.objects.create(
        board=board_obj, stage=todo, title="On time", due_date=today + timedelta(days=2)
    )

    content = auth_client.get(reverse("kanban")).content.decode()

    assert "kanban-count-overdue" not in content
