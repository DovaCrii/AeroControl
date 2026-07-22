import csv
import json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import ImportBatch
from apps.registry.models import Aircraft, CostCenter, Operator


SCHEMAS = {
    "cost_centers": (CostCenter, ["code", "name"]),
    "aircraft": (Aircraft, ["registration", "type", "model", "manufacturer", "year", "cost_center", "status"]),
    "operators": (Operator, ["employee_id", "full_name", "email", "phone", "cost_center"]),
}


class Command(BaseCommand):
    help = "Validate and optionally apply Chapter 1 CSV sources using chapter1-v1."

    def add_arguments(self, parser):
        for name in SCHEMAS:
            parser.add_argument(f"--{name.replace('_', '-')}", type=Path)
        parser.add_argument("--apply", action="store_true")
        parser.add_argument("--workbook", type=Path, help="Excel workbook with cost_centers, aircraft and operators sheets.")
        parser.add_argument("--json", action="store_true", dest="as_json")

    def parse_file(self, key, path):
        if not path or not path.is_file():
            raise CommandError(f"Missing source for {key}: {path}")
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            expected = SCHEMAS[key][1]
            if reader.fieldnames != expected:
                raise CommandError(f"{key} columns must be: {','.join(expected)}")
            rows = [{field: (value or "").strip() for field, value in row.items()} for row in reader]
        if len(rows) > 5000:
            raise CommandError(f"{key} exceeds the 5000-row limit")
        unique_field = {"cost_centers": "code", "aircraft": "registration", "operators": "employee_id"}[key]
        values = [row[unique_field] for row in rows]
        if any(not value for value in values) or len(values) != len(set(values)):
            raise CommandError(f"{key} contains empty or duplicate {unique_field} values")
        model = SCHEMAS[key][0]
        if model.objects.filter(**{f"{unique_field}__in": values}).exists():
            raise CommandError(f"{key} contains values already present in the database")
        if key in {"aircraft", "operators"}:
            for row in rows:
                if not CostCenter.objects.filter(code=row["cost_center"], is_active=True).exists():
                    raise CommandError(f"{key} references unknown cost center {row['cost_center']}")
        if key == "aircraft":
            for row in rows:
                row["year"] = int(row["year"]) if row["year"].isdigit() else None
                row["status"] = row["status"] or "active"
        return rows

    def parse_workbook(self, path):
        if not path or not path.is_file():
            raise CommandError(f"Missing workbook: {path}")
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        payload = {}
        sheet_names = {"cost_centers": "cost_centers", "aircraft": "aircraft", "operators": "operators"}
        for key, sheet_name in sheet_names.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))]
            expected = SCHEMAS[key][1]
            if headers != expected:
                raise CommandError(f"{sheet_name} columns must be: {','.join(expected)}")
            rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({field: str(value or "").strip() for field, value in zip(headers, values)})
            payload[key] = rows
        if not payload:
            raise CommandError("Workbook must contain at least one supported sheet.")
        return payload

    def handle(self, *args, **options):
        paths = {key: options.get(key) for key in SCHEMAS if options.get(key)}
        if not paths and not options.get("workbook"):
            raise CommandError("Provide at least one Chapter 1 CSV source.")
        payload = self.parse_workbook(options["workbook"]) if options.get("workbook") else {key: self.parse_file(key, path) for key, path in paths.items()}
        if options.get("workbook"):
            for key, rows in payload.items():
                # Reuse CSV validation by applying the same invariants in-memory.
                if len(rows) > 5000:
                    raise CommandError(f"{key} exceeds the 5000-row limit")
                unique_field = {"cost_centers": "code", "aircraft": "registration", "operators": "employee_id"}[key]
                values = [row.get(unique_field, "") for row in rows]
                if any(not value for value in values) or len(values) != len(set(values)):
                    raise CommandError(f"{key} contains empty or duplicate {unique_field} values")
                if key == "aircraft":
                    for row in rows:
                        row["year"] = int(row["year"]) if row.get("year", "").isdigit() else None
                        row["status"] = row.get("status") or "active"
                if key in {"aircraft", "operators"}:
                    for row in rows:
                        if not CostCenter.objects.filter(code=row.get("cost_center"), is_active=True).exists():
                            raise CommandError(f"{key} references unknown cost center {row.get('cost_center')}")
        result = {"mapping": "chapter1-v1", "apply": options["apply"], "rows": {key: len(rows) for key, rows in payload.items()}}
        if options["apply"]:
            with transaction.atomic():
                for key, rows in payload.items():
                    model = SCHEMAS[key][0]
                    batch = ImportBatch.objects.create(actor=None, entity=f"chapter1.{key}", rows=rows)
                    created = []
                    for row in rows:
                        if key == "aircraft":
                            row["cost_center_id"] = CostCenter.objects.get(code=row.pop("cost_center")).pk
                        elif key == "operators":
                            row["cost_center_id"] = CostCenter.objects.get(code=row.pop("cost_center")).pk
                        created.append(model.objects.create(**row))
                    batch.created_ids = [str(obj.pk) for obj in created]
                    batch.save(update_fields=["created_ids", "updated_at"])
        if options["as_json"]:
            self.stdout.write(json.dumps(result, ensure_ascii=False))
        else:
            self.stdout.write(f"mapping: {result['mapping']}")
            self.stdout.write(f"apply: {result['apply']}")
            for key, count in result["rows"].items():
                self.stdout.write(f"{key}: {count}")
